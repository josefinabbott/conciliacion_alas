#!/usr/bin/env python3
"""
Conciliación mensual de cobros de Alas (courier "alas flapp") vs. tarifa correcta Flapp.

Uso:
    python3 reconciliar_alas.py <archivo_entrada.csv> <mes_label> [carpeta_salida]

Ejemplo:
    python3 reconciliar_alas.py "Revision Pedidos Alas Agosto.csv" "Agosto 2026" ./salida

Qué hace:
  1. Lee el excel/csv que junta el cobro real de Alas ("Cobrado x Alas") con los datos
     del pedido en Flapp (columnas general.*, buyer.*, client.*, courier.*).
  2. Recalcula, de forma independiente, cuál DEBERÍA ser la tarifa según la Matriz de
     Tarifas Flapp vigente (mismo motor usado para auditar y corregir ManualQuotes:
     tienda XS para Farmacia Bosques/Musse Cosmetics, y la fórmula de zona/base para
     el resto de los clientes).
  3. Triangula 3 números por cada envío:
       - tarifa_correcta         -> lo que debería costar según la matriz vigente
       - courier.shippingFee     -> lo que Flapp tenía cargado/esperado para ese envío
       - Cobrado x Alas          -> lo que Alas facturó realmente
     y clasifica cada fila en una de estas categorías:
       - sin_discrepancia         : los 3 valores coinciden (o difieren por <$1)
       - cancelado_sin_cargo      : envío cancelado, Alas no cobró nada (no es error)
       - manualquote_desactualizada: Alas cobró lo correcto, pero Flapp tenía cargada
                                     una tarifa vieja/errónea (no hay nada que reclamarle
                                     a Alas; hay que actualizar ManualQuotes)
       - alas_cobro_distinto      : Flapp tenía la tarifa correcta cargada, pero Alas
                                     cobró un monto distinto -> ESTO es lo que hay que
                                     disputarle a Alas
       - revisar_manualmente      : ambos valores difieren de la tarifa correcta Y entre
                                     sí, en un envío no cancelado -> requiere revisión
                                     manual (posible cambio real de zona/bodega ese mes)
  4. Genera un Excel de varias hojas con el detalle, y un borrador de email para Finanzas.

Requisitos: pandas, openpyxl (ya instalados en el ambiente de Claude).
"""
import sys
import os
import csv
import json
import math
from collections import defaultdict, Counter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RATES_PATH = os.path.join(SCRIPT_DIR, "..", "assets", "rates_table.json")

XS_CLIENTS = {"Farmacia Bosques", "Musse Cosmetics"}
UMBRAL_OK = 1  # diferencias menores a $1 se consideran "sin discrepancia" (redondeo)


def cargar_matriz_tarifas():
    with open(RATES_PATH, encoding="utf-8") as f:
        rates = json.load(f)
    by_commune = {r["commune"].strip().lower(): r for r in rates}
    return by_commune


def tarifa_correcta(by_commune, cliente, origen, destino):
    """Reproduce el motor de tarifas 'alas flapp' validado contra Metabase (question 2418)."""
    origen = (origen or "").strip().lower()
    destino = (destino or "").strip().lower()
    if origen not in by_commune or destino not in by_commune:
        return None, f"comuna_no_encontrada(origen={origen!r}, destino={destino!r})"
    ro = by_commune[origen]
    rd = by_commune[destino]
    if cliente in XS_CLIENTS:
        return math.ceil(rd["xs"] * 1.19 - 0.000001), None
    if ro["zona"] == "RM" and rd["zona"] == "RM" and rd["categoria"] == "Urbano":
        base = rd["rm_a_reg"]
    elif ro["base_code"] == rd["base_code"]:
        base = rd["base_a_base"]
    elif ro["zona"] == "RM":
        base = rd["rm_a_reg"]
    else:
        base = rd["reg_a_reg"]
    return round(base * 1.19), None


def to_num(x):
    if x is None:
        return None
    x = str(x).strip()
    if x == "":
        return None
    try:
        return float(x)
    except ValueError:
        return None


COLUMNAS_ESPERADAS = [
    "general.clientName", "general.localName", "general.orderId", "general.orderCode",
    "general.shipmentId", "general.shipmentExternalId", "general.shipmentStatus",
    "general.closedDate", "general.commune", "general.destinationCommune",
    "courier.shippingFee", "Cobrado x Alas",
]


def leer_filas(input_path):
    """Acepta .csv o .xlsx/.xls indistintamente y devuelve (rows, fieldnames)."""
    ext = os.path.splitext(input_path)[1].lower()
    if ext in (".xlsx", ".xls", ".xlsm"):
        import pandas as pd
        df = pd.read_excel(input_path, dtype=str)
        df = df.where(df.notna(), "")
        rows = df.to_dict(orient="records")
        fieldnames = list(df.columns)
        return rows, fieldnames
    else:
        with open(input_path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            fieldnames = reader.fieldnames
        return rows, fieldnames


def procesar(input_path):
    rows, fieldnames = leer_filas(input_path)

    faltantes = [c for c in COLUMNAS_ESPERADAS if c not in fieldnames]
    if faltantes:
        raise SystemExit(
            "El archivo de entrada no tiene las columnas esperadas. Faltan: "
            + ", ".join(faltantes)
            + "\nColumnas encontradas: " + ", ".join(fieldnames)
        )

    by_commune = cargar_matriz_tarifas()

    resultados = []
    for r in rows:
        cliente = r.get("general.clientName", "")
        origen = r.get("general.commune", "")
        destino = r.get("general.destinationCommune", "")
        cobrado = to_num(r.get("Cobrado x Alas"))
        cargado = to_num(r.get("courier.shippingFee"))
        estado = r.get("general.shipmentStatus", "")

        correcta, error_comuna = tarifa_correcta(by_commune, cliente, origen, destino)

        fila = {
            "shipmentId": r.get("general.shipmentId", ""),
            "orderId": r.get("general.orderId", ""),
            "orderCode": r.get("general.orderCode", ""),
            "shipmentExternalId": r.get("general.shipmentExternalId", ""),
            "cliente": cliente,
            "local": r.get("general.localName", ""),
            "origen": origen,
            "destino": destino,
            "fecha_cierre": r.get("general.closedDate", ""),
            "estado_envio": estado,
            "tarifa_correcta_matriz": correcta,
            "cargado_en_flapp": cargado,
            "cobrado_por_alas": cobrado,
        }

        if error_comuna:
            fila["categoria"] = "comuna_no_reconocida"
            fila["diferencia_alas_vs_correcta"] = None
            fila["nota"] = error_comuna
            resultados.append(fila)
            continue

        d_alas = cobrado is not None and abs(cobrado - correcta) > UMBRAL_OK
        d_flapp = cargado is not None and abs(cargado - correcta) > UMBRAL_OK
        diff_alas_correcta = (cobrado - correcta) if cobrado is not None else None

        if estado == "cancelled" and (cobrado or 0) == 0 and (cargado or 0) == 0:
            categoria = "cancelado_sin_cargo"
        elif not d_alas and not d_flapp:
            categoria = "sin_discrepancia"
        elif d_flapp and not d_alas:
            categoria = "manualquote_desactualizada"
        elif d_alas and not d_flapp:
            categoria = "alas_cobro_distinto"
        else:
            categoria = "revisar_manualmente"

        fila["categoria"] = categoria
        fila["diferencia_alas_vs_correcta"] = diff_alas_correcta
        fila["nota"] = ""
        resultados.append(fila)

    return resultados


def construir_excel(resultados, mes_label, out_path):
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(resultados)

    orden_categorias = [
        "alas_cobro_distinto",
        "revisar_manualmente",
        "manualquote_desactualizada",
        "comuna_no_reconocida",
        "cancelado_sin_cargo",
        "sin_discrepancia",
    ]
    df["_orden"] = df["categoria"].apply(lambda c: orden_categorias.index(c) if c in orden_categorias else 99)
    df = df.sort_values(["_orden", "diferencia_alas_vs_correcta"], key=lambda s: s if s.name == "_orden" else s.abs(), ascending=[True, False]).drop(columns="_orden")

    disputas = df[df["categoria"] == "alas_cobro_distinto"].copy()
    revisar = df[df["categoria"] == "revisar_manualmente"].copy()
    desactualizadas = df[df["categoria"] == "manualquote_desactualizada"].copy()
    sin_comuna = df[df["categoria"] == "comuna_no_reconocida"].copy()

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Hoja Detalle completo primero (para que las formulas de Resumen puedan apuntar a ella)
        df.to_excel(writer, sheet_name="Detalle completo", index=False)
        disputas.to_excel(writer, sheet_name="Disputas con Alas", index=False)
        revisar.to_excel(writer, sheet_name="Revisar manualmente", index=False)
        desactualizadas.to_excel(writer, sheet_name="ManualQuotes desact.", index=False)
        if len(sin_comuna):
            sin_comuna.to_excel(writer, sheet_name="Comunas no reconocidas", index=False)

        wb = writer.book
        detalle_ws = wb["Detalle completo"]
        n_rows = len(df) + 1  # + header

        # ---- Hoja Resumen (con formulas SUMIFS/COUNTIFS sobre Detalle completo) ----
        resumen_ws = wb.create_sheet("Resumen Ejecutivo", 0)
        resumen_ws["A1"] = f"Conciliación de cobros Alas — {mes_label}"
        resumen_ws["A1"].font = Font(name="Arial", size=14, bold=True)
        resumen_ws["A2"] = "Generado automáticamente comparando: tarifa correcta según Matriz de Tarifas Flapp vs. lo cargado en Flapp vs. lo cobrado por Alas."
        resumen_ws["A2"].font = Font(name="Arial", size=9, italic=True)

        headers = ["Categoría", "Cantidad de envíos", "Suma diferencia (Alas - correcta) $", "Descripción"]
        desc = {
            "alas_cobro_distinto": "Flapp tenía la tarifa correcta cargada, pero Alas cobró un monto distinto. REQUIERE DISPUTA con Alas.",
            "revisar_manualmente": "Tanto lo cargado en Flapp como lo cobrado por Alas difieren de la tarifa correcta, y entre sí. Revisar caso a caso (posible cambio real de zona/bodega).",
            "manualquote_desactualizada": "Alas cobró correctamente, pero Flapp tenía cargada una tarifa vieja/errónea. No hay nada que reclamarle a Alas; actualizar ManualQuotes.",
            "comuna_no_reconocida": "No se pudo calcular la tarifa correcta (comuna no está en la matriz de tarifas). Revisar manualmente.",
            "cancelado_sin_cargo": "Envío cancelado, Alas no cobró. Informativo, no requiere acción.",
            "sin_discrepancia": "Todo cuadra: tarifa correcta = cargado en Flapp = cobrado por Alas.",
        }
        row0 = 4
        for j, h in enumerate(headers, start=1):
            c = resumen_ws.cell(row=row0, column=j, value=h)
            c.font = Font(name="Arial", bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="4472C4")
        # Verificar posición real de las columnas 'categoria' y 'diferencia_alas_vs_correcta'
        cols = list(df.columns)
        col_cat = get_column_letter(cols.index("categoria") + 1)
        col_dif = get_column_letter(cols.index("diferencia_alas_vs_correcta") + 1)

        r = row0
        for cat in orden_categorias:
            r += 1
            resumen_ws.cell(row=r, column=1, value=cat)
            resumen_ws.cell(row=r, column=2, value=f"=COUNTIF('Detalle completo'!{col_cat}2:{col_cat}{n_rows},A{r})")
            resumen_ws.cell(row=r, column=3, value=f"=SUMIF('Detalle completo'!{col_cat}2:{col_cat}{n_rows},A{r},'Detalle completo'!{col_dif}2:{col_dif}{n_rows})")
            resumen_ws.cell(row=r, column=4, value=desc.get(cat, ""))
        r += 1
        resumen_ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
        resumen_ws.cell(row=r, column=2, value=f"=SUM(B{row0+1}:B{r-1})").font = Font(bold=True)
        resumen_ws.cell(row=r, column=3, value=f"=SUM(C{row0+1}:C{r-1})").font = Font(bold=True)

        for col, width in zip("ABCD", (28, 20, 32, 90)):
            resumen_ws.column_dimensions[col].width = width
        for row in resumen_ws.iter_rows(min_row=row0+1, max_row=r):
            for cell in row:
                cell.font = Font(name="Arial", bold=cell.font.bold)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Formato de todas las hojas de datos: fuente Arial, encabezado en negrita, columnas auto
        for sheet_name in ["Detalle completo", "Disputas con Alas", "Revisar manualmente", "ManualQuotes desact.", "Comunas no reconocidas"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for cell in ws[1]:
                cell.font = Font(name="Arial", bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4472C4")
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(name="Arial")
            for i, col_cells in enumerate(ws.columns, start=1):
                max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
                ws.column_dimensions[get_column_letter(i)].width = min(max(10, max_len + 2), 45)

        wb.move_sheet("Resumen Ejecutivo", offset=-(len(wb.sheetnames)-1))

    return df, disputas, revisar, desactualizadas, sin_comuna


def construir_email(mes_label, disputas, revisar, desactualizadas, sin_comuna):
    total_disputa = disputas["diferencia_alas_vs_correcta"].sum() if len(disputas) else 0
    n_disputa = len(disputas)
    n_revisar = len(revisar)
    n_desact = len(desactualizadas)

    top = disputas.reindex(disputas["diferencia_alas_vs_correcta"].abs().sort_values(ascending=False).index).head(10)

    lineas_top = []
    for _, row in top.iterrows():
        signo = "cobró de más" if row["diferencia_alas_vs_correcta"] > 0 else "cobró de menos"
        lineas_top.append(
            f"  - Envío {row['shipmentId']} ({row['origen']} -> {row['destino']}, cliente {row['cliente']}): "
            f"tarifa correcta ${row['tarifa_correcta_matriz']:,.0f}, Alas cobró ${row['cobrado_por_alas']:,.0f} "
            f"({signo} ${abs(row['diferencia_alas_vs_correcta']):,.0f})"
        )
    lineas_top_txt = "\n".join(lineas_top) if lineas_top else "  (sin casos)"

    cuerpo = f"""Asunto: Conciliación cobros Alas — {mes_label} — {n_disputa} envíos en disputa

Hola,

Adjunto la conciliación de cobros de Alas correspondiente a {mes_label}. Comparamos, para cada
envío, la tarifa que debería aplicar según nuestra matriz de tarifas vigente, lo que teníamos
cargado en Flapp, y lo que Alas efectivamente cobró.

Resumen:
  - {n_disputa} envíos donde Alas cobró un monto distinto al que corresponde según la tarifa
    acordada (nuestra carga en Flapp era correcta). Diferencia neta: ${total_disputa:,.0f}
    ({"a favor nuestro" if total_disputa < 0 else "en contra nuestra" if total_disputa > 0 else "neutro"}).
  - {n_revisar} envíos requieren revisión manual (tanto lo cargado como lo cobrado difieren de
    la tarifa correcta, y entre sí — posible cambio real de zona o bodega ese mes).
  - {n_desact} envíos donde Alas cobró correctamente pero teníamos una tarifa vieja cargada en
    ManualQuotes (no es un problema de facturación de Alas; ya lo estamos corrigiendo
    internamente).

Los 10 casos de mayor impacto a disputar con Alas:
{lineas_top_txt}

Se adjunta el detalle completo en Excel (hoja "Disputas con Alas") con el desglose envío por
envío para que puedan revisarlo con Alas directamente.

Quedo atento a comentarios antes de enviarlo.

Saludos,
"""
    return cuerpo


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    input_path = sys.argv[1]
    mes_label = sys.argv[2]
    out_dir = sys.argv[3] if len(sys.argv) > 3 else "."
    os.makedirs(out_dir, exist_ok=True)

    resultados = procesar(input_path)

    slug = mes_label.replace(" ", "_")
    xlsx_path = os.path.join(out_dir, f"Conciliacion_Alas_{slug}.xlsx")
    df, disputas, revisar, desactualizadas, sin_comuna = construir_excel(resultados, mes_label, xlsx_path)

    email_txt = construir_email(mes_label, disputas, revisar, desactualizadas, sin_comuna)
    email_path = os.path.join(out_dir, f"Borrador_Email_Alas_{slug}.md")
    with open(email_path, "w", encoding="utf-8") as f:
        f.write(email_txt)

    print(f"OK. Total filas procesadas: {len(resultados)}")
    print(Counter(r["categoria"] for r in resultados))
    print(f"Excel: {xlsx_path}")
    print(f"Email: {email_path}")


if __name__ == "__main__":
    main()
