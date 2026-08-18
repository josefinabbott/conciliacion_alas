#!/usr/bin/env python3
"""
Conciliación mensual de cobros de Alas (courier "alas flapp") vs. tarifa correcta Flapp.

Uso (modo 1 -- un archivo ya cruzado por finanzas, con columna "Cobrado x Alas"):
    python3 reconciliar_alas.py <archivo_entrada.csv> <mes_label> [carpeta_salida]

Uso (modo 2 -- dos archivos crudos, sin cruzar a mano: el export de Flapp + la factura
cruda de Alas -- RECOMENDADO, evita el cruce manual de finanzas):
    python3 reconciliar_alas.py <export_flapp.csv> <mes_label> [carpeta_salida] --alas <factura_alas.csv>

Ejemplo:
    python3 reconciliar_alas.py "Revision Pedidos Alas Agosto.csv" "Agosto 2026" ./salida
    python3 reconciliar_alas.py "export_flapp_agosto.csv" "Agosto 2026" ./salida --alas "factura_alas_agosto.csv"

Qué hace:
  1. Lee el excel/csv que junta el cobro real de Alas ("Cobrado x Alas") con los datos
     del pedido en Flapp (columnas general.*, buyer.*, client.*, courier.*).
  2. Recalcula, de forma independiente, cuál DEBERÍA ser la tarifa según la Matriz de
     Tarifas Flapp vigente (mismo motor usado para auditar y corregir ManualQuotes:
     tienda XS para Farmacia Bosques/Musse Cosmetics, y la fórmula de zona/base para
     el resto de los clientes).
  3. Triangula 3 números por cada envío:
       - tarifa_correcta         -> lo que debería costar según la matriz vigente
       - courier.shippingFee     -> lo que Flapp tenía cargado/esperado para ese envío
       - Cobrado x Alas          -> lo que Alas facturó realmente
     y clasifica cada fila en una de estas categorías:
       - sin_discrepancia         : los 3 valores coinciden (o difieren por <$1)
       - cancelado_sin_cargo      : envío cancelado, Alas no cobró nada (no es error)
       - manualquote_desactualizada: Alas cobró lo correcto, pero Flapp tenía cargada
                                     una tarifa vieja/errónea (no hay nada que reclamarle
                                     a Alas; hay que actualizar ManualQuotes)
       - alas_cobro_distinto      : Flapp tenía la tarifa correcta cargada, pero Alas
                                     cobró un monto distinto -> ESTO es lo que hay que
                                     disputarle a Alas
       - revisar_manualmente      : ambos valores difieren de la tarifa correcta Y entre
                                     sí, en un envío no cancelado -> requiere revisión
                                     manual (posible cambio real de zona/bodega ese mes)
  4. Genera un Excel de varias hojas con el detalle, y un borrador de email para Finanzas.

Requisitos: pandas, openpyxl (ya instalados en el ambiente de Claude).
"""
import sys
import os
import csv
import json
import math
import unicodedata
from collections import defaultdict, Counter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RATES_PATH = os.path.join(SCRIPT_DIR, "..", "assets", "rates_table.json")

XS_CLIENTS = {"Farmacia Bosques", "Musse Cosmetics"}
UMBRAL_OK = 1  # diferencias menores a $1 se consideran "sin discrepancia" (redondeo)

# Alias de nombres de comuna que aparecen en archivos reales de Alas/Flapp pero no calzan
# textualmente con el nombre "canónico" en la matriz de tarifas (abreviaturas, sufijos, etc.).
# Se aplican DESPUÉS de sacar tildes y pasar a minúscula.
COMUNA_ALIASES = {
    "santiago": "santiago centro",
    "calera": "la calera",
    "natales": "puerto natales",
    "san vicente": "san vicente de tagua tagua",
    "cabo de hornos (ex. navarino)": "cabo de hornos",
}


def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def normalizar_comuna(nombre):
    n = strip_accents((nombre or "").strip().lower())
    return COMUNA_ALIASES.get(n, n)


def cargar_matriz_tarifas():
    with open(RATES_PATH, encoding="utf-8") as f:
        rates = json.load(f)
    # Se indexa tanto por el nombre tal cual viene en la matriz como por su versión sin tildes,
    # para que comunas escritas con o sin tilde (ej. "Los Ángeles" / "los angeles") calcen igual.
    by_commune = {}
    for r in rates:
        clave = strip_accents(r["commune"].strip().lower())
        by_commune[clave] = r
    return by_commune


def tarifa_correcta(by_commune, cliente, origen, destino):
    """Reproduce el motor de tarifas 'alas flapp' validado contra Metabase (question 2418)."""
    origen = normalizar_comuna(origen)
    destino = normalizar_comuna(destino)
    if origen not in by_commune or destino not in by_commune:
        return None, f"comuna_no_encontrada(origen={origen!r}, destino={destino!r})"
    ro = by_commune[origen]
    rd = by_commune[destino]
    if cliente in XS_CLIENTS:
        return math.ceil(rd["xs"] * 1.19 - 0.000001), None
    if ro["zona"] == "RM" and rd["zona"] == "RM" and rd["categoria"] == "Urbano":
        base = rd["rm_a_reg"]
    elif ro["base_code"] == rd["base_code"]:
        base = rd["base_a_base"]
    elif ro["zona"] == "RM":
        base = rd["rm_a_reg"]
    else:
        base = rd["reg_a_reg"]
    return round(base * 1.19), None


def to_num(x):
    if x is None:
        return None
    x = str(x).strip()
    if x == "":
        return None
    try:
        return float(x)
    except ValueError:
        return None


def parse_moneda(x):
    """Convierte strings de plata tipo '$4,190' o '4.190' a número. None si no se puede."""
    if x is None:
        return None
    s = str(x).strip()
    if s == "":
        return None
    s = s.replace("$", "").replace(" ", "")
    # Formato chileno: "," o "." como separador de miles, sin decimales.
    s = s.replace(",", "").replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


def normalizar_cliente(nombre):
    """Corrige variantes/typos comunes de nombre de cliente (ej. 'Farmacias Bosques' -> 'Farmacia Bosques')."""
    n = (nombre or "").strip().lower()
    if "farmacia" in n and "bosque" in n:
        return "Farmacia Bosques"
    if "musse" in n:
        return "Musse Cosmetics"
    return (nombre or "").strip()


def clasificar(cliente, origen, destino, cobrado, cargado, estado, by_commune):
    """Lógica de clasificación compartida entre el modo de 1 archivo y el de 2 archivos.
    Devuelve (categoria, tarifa_correcta, diferencia, nota)."""
    correcta, error_comuna = tarifa_correcta(by_commune, cliente, origen, destino)
    if error_comuna:
        return "comuna_no_reconocida", None, None, error_comuna

    d_alas = cobrado is not None and abs(cobrado - correcta) > UMBRAL_OK
    d_flapp = cargado is not None and abs(cargado - correcta) > UMBRAL_OK
    diff_alas_correcta = (cobrado - correcta) if cobrado is not None else None

    if estado == "cancelled" and (cobrado or 0) == 0 and (cargado or 0) == 0:
        categoria = "cancelado_sin_cargo"
    elif not d_alas and not d_flapp:
        categoria = "sin_discrepancia"
    elif d_flapp and not d_alas:
        categoria = "manualquote_desactualizada"
    elif d_alas and not d_flapp:
        categoria = "alas_cobro_distinto"
    else:
        categoria = "revisar_manualmente"

    return categoria, correcta, diff_alas_correcta, ""


COLUMNAS_ESPERADAS = [
    "general.clientName", "general.localName", "general.orderId", "general.orderCode",
    "general.shipmentId", "general.shipmentExternalId", "general.shipmentStatus",
    "general.closedDate", "general.commune", "general.destinationCommune",
    "courier.shippingFee", "Cobrado x Alas",
]

# El export "crudo" de Flapp (sin cruzar con Alas) trae las mismas columnas de arriba
# menos "Cobrado x Alas" (que es justo lo que se agrega cruzando con el archivo de Alas).
COLUMNAS_FLAPP_BRUTO = [c for c in COLUMNAS_ESPERADAS if c != "Cobrado x Alas"]


def leer_filas(input_path):
    """Acepta .csv o .xlsx/.xls indistintamente y devuelve (rows, fieldnames)."""
    ext = os.path.splitext(input_path)[1].lower()
    if ext in (".xlsx", ".xls", ".xlsm"):
        import pandas as pd
        df = pd.read_excel(input_path, dtype=str)
        df = df.where(df.notna(), "")
        rows = df.to_dict(orient="records")
        fieldnames = list(df.columns)
        return rows, fieldnames
    else:
        with open(input_path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            fieldnames = reader.fieldnames
        return rows, fieldnames


def procesar(input_path):
    rows, fieldnames = leer_filas(input_path)

    faltantes = [c for c in COLUMNAS_ESPERADAS if c not in fieldnames]
    if faltantes:
        raise SystemExit(
            "El archivo de entrada no tiene las columnas esperadas. Faltan: "
            + ", ".join(faltantes)
            + "\nColumnas encontradas: " + ", ".join(fieldnames)
        )

    by_commune = cargar_matriz_tarifas()

    resultados = []
    for r in rows:
        cliente = r.get("general.clientName", "")
        origen = r.get("general.commune", "")
        destino = r.get("general.destinationCommune", "")
        cobrado = to_num(r.get("Cobrado x Alas"))
        cargado = to_num(r.get("courier.shippingFee"))
        estado = r.get("general.shipmentStatus", "")

        categoria, correcta, diff_alas_correcta, nota = clasificar(
            cliente, origen, destino, cobrado, cargado, estado, by_commune
        )

        fila = {
            "shipmentId": r.get("general.shipmentId", ""),
            "orderId": r.get("general.orderId", ""),
            "orderCode": r.get("general.orderCode", ""),
            "shipmentExternalId": r.get("general.shipmentExternalId", ""),
            "cliente": cliente,
            "local": r.get("general.localName", ""),
            "origen": origen,
            "destino": destino,
            "fecha_cierre": r.get("general.closedDate", ""),
            "estado_envio": estado,
            "tarifa_correcta_matriz": correcta,
            "cargado_en_flapp": cargado,
            "cobrado_por_alas": cobrado,
            "categoria": categoria,
            "diferencia_alas_vs_correcta": diff_alas_correcta,
            "nota": nota,
        }
        resultados.append(fila)

    return resultados


def leer_flapp_bruto(path):
    """Lee el export crudo de Flapp (el que se descarga directo del sistema, sin cruzar con
    Alas): separador ';', columnas general.*/buyer.*/client.*/courier.* Devuelve un dict
    indexado por general.shipmentId."""
    with open(path, encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.DictReader(f, delimiter=delimiter)
        rows = list(reader)
        fieldnames = reader.fieldnames

    faltantes = [c for c in COLUMNAS_FLAPP_BRUTO if c not in (fieldnames or [])]
    if faltantes:
        raise SystemExit(
            "El archivo de Flapp no tiene las columnas esperadas. Faltan: " + ", ".join(faltantes)
            + "\nColumnas encontradas: " + ", ".join(fieldnames or [])
        )

    by_id = {}
    for r in rows:
        sid = str(r.get("general.shipmentId", "")).strip()
        if sid:
            by_id[sid] = r
    return by_id


ALAS_COL_SHIPMENT_ID = 1
ALAS_COL_CLIENTE = 2
ALAS_COL_COMUNA_ORIGEN = 4
ALAS_COL_FOLIO = 7
ALAS_COL_COMUNA_DESTINO = 16
ALAS_COL_TARIFA = 23
ALAS_COL_MODELO_TARIFARIO = 24
ALAS_MIN_COLUMNAS = 25


def leer_alas_bruto(path):
    """Lee el excel/csv crudo que manda Alas con su propia facturación: trae filas de título y
    un encabezado repetido cada cierta cantidad de páginas, así que se identifica la fila de
    encabezado real buscando 'CLIENTE' y 'Tarifa', y se descartan filas basura (todo lo que no
    empiece con el marcador 'Flapp' en la primera columna)."""
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        all_rows = list(reader)

    header_idx = None
    for i, r in enumerate(all_rows):
        if "CLIENTE" in r and "Tarifa" in r:
            header_idx = i
            break
    if header_idx is None:
        raise SystemExit(
            "No se encontró la fila de encabezado (con 'CLIENTE' y 'Tarifa') en el archivo de Alas. "
            "¿Es el formato correcto?"
        )

    filas = []
    for r in all_rows[header_idx + 1:]:
        if len(r) < ALAS_MIN_COLUMNAS:
            continue
        if r[0].strip() != "Flapp":
            continue  # fila basura o encabezado repetido
        filas.append(r)
    return filas


def procesar_dos_archivos(path_flapp, path_alas):
    """Concilia a partir de los DOS archivos crudos (el export de Flapp + la factura cruda de
    Alas), sin necesitar que finanzas los cruce a mano de antemano. El cruce se hace por
    general.shipmentId (que en el archivo de Alas es la columna sin nombre justo después de
    'B2B', y coincide con el prefijo del 'Folio')."""
    flapp_by_id = leer_flapp_bruto(path_flapp)
    alas_filas = leer_alas_bruto(path_alas)
    by_commune = cargar_matriz_tarifas()

    resultados = []
    matched_ids = set()

    for r in alas_filas:
        sid = r[ALAS_COL_SHIPMENT_ID].strip()
        cliente_raw = r[ALAS_COL_CLIENTE]
        cliente = normalizar_cliente(cliente_raw)
        origen = r[ALAS_COL_COMUNA_ORIGEN]
        destino = r[ALAS_COL_COMUNA_DESTINO]
        tarifa_alas_bruta = parse_moneda(r[ALAS_COL_TARIFA])
        # La columna "Tarifa" del archivo crudo de Alas viene SIN IVA (validado contra la
        # matriz: p.ej. para Musse/XS coincide exacto con el valor "xs" bruto de la matriz, y
        # para el resto coincide con "base_a_base"/"reg_a_reg" brutos) -- hay que agregarle el
        # 19% para que sea comparable con "tarifa_correcta_matriz" y "courier.shippingFee",
        # que sí vienen con IVA incluido.
        tarifa_alas = round(tarifa_alas_bruta * 1.19) if tarifa_alas_bruta is not None else None
        folio = r[ALAS_COL_FOLIO]
        modelo_tarifario_alas = r[ALAS_COL_MODELO_TARIFARIO]

        flapp_row = flapp_by_id.get(sid)

        if flapp_row is None:
            resultados.append({
                "shipmentId": sid,
                "orderId": "",
                "orderCode": folio,
                "shipmentExternalId": "",
                "cliente": cliente,
                "local": "",
                "origen": origen,
                "destino": destino,
                "fecha_cierre": "",
                "estado_envio": "",
                "tarifa_correcta_matriz": None,
                "cargado_en_flapp": None,
                "cobrado_por_alas": tarifa_alas,
                "categoria": "sin_dato_flapp",
                "diferencia_alas_vs_correcta": None,
                "nota": f"Alas facturó este envío (folio {folio}) pero no aparece en el export de Flapp -- revisar manualmente.",
            })
            continue

        matched_ids.add(sid)
        cargado = to_num(flapp_row.get("courier.shippingFee"))
        estado = flapp_row.get("general.shipmentStatus", "")

        categoria, correcta, diff_alas_correcta, nota = clasificar(
            cliente, origen, destino, tarifa_alas, cargado, estado, by_commune
        )
        if categoria != "comuna_no_reconocida" and modelo_tarifario_alas:
            nota = (nota + " " if nota else "") + f"[Modelo tarifario usado por Alas: {modelo_tarifario_alas}]"

        resultados.append({
            "shipmentId": sid,
            "orderId": flapp_row.get("general.orderId", ""),
            "orderCode": flapp_row.get("general.orderCode", ""),
            "shipmentExternalId": flapp_row.get("general.shipmentExternalId", ""),
            "cliente": cliente,
            "local": flapp_row.get("general.localName", ""),
            "origen": origen,
            "destino": destino,
            "fecha_cierre": flapp_row.get("general.closedDate", ""),
            "estado_envio": estado,
            "tarifa_correcta_matriz": correcta,
            "cargado_en_flapp": cargado,
            "cobrado_por_alas": tarifa_alas,
            "categoria": categoria,
            "diferencia_alas_vs_correcta": diff_alas_correcta,
            "nota": nota,
        })

    # Envíos que están en el export de Flapp pero Alas no facturó (todavía, o nunca) --
    # no se puede reconciliar sin la factura, se dejan marcados aparte.
    for sid, flapp_row in flapp_by_id.items():
        if sid in matched_ids:
            continue
        resultados.append({
            "shipmentId": sid,
            "orderId": flapp_row.get("general.orderId", ""),
            "orderCode": flapp_row.get("general.orderCode", ""),
            "shipmentExternalId": flapp_row.get("general.shipmentExternalId", ""),
            "cliente": normalizar_cliente(flapp_row.get("general.clientName", "")),
            "local": flapp_row.get("general.localName", ""),
            "origen": flapp_row.get("general.commune", ""),
            "destino": flapp_row.get("general.destinationCommune", ""),
            "fecha_cierre": flapp_row.get("general.closedDate", ""),
            "estado_envio": flapp_row.get("general.shipmentStatus", ""),
            "tarifa_correcta_matriz": None,
            "cargado_en_flapp": to_num(flapp_row.get("courier.shippingFee")),
            "cobrado_por_alas": None,
            "categoria": "sin_dato_alas",
            "diferencia_alas_vs_correcta": None,
            "nota": "Este envío está en el export de Flapp pero Alas no lo facturó en este archivo -- puede estar pendiente o fuera del período.",
        })

    return resultados


def construir_excel(resultados, mes_label, out_path):
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(resultados)

    orden_categorias = [
        "alas_cobro_distinto",
        "revisar_manualmente",
        "sin_dato_flapp",
        "sin_dato_alas",
        "manualquote_desactualizada",
        "comuna_no_reconocida",
        "cancelado_sin_cargo",
        "sin_discrepancia",
    ]
    df["_orden"] = df["categoria"].apply(lambda c: orden_categorias.index(c) if c in orden_categorias else 99)
    df = df.sort_values(["_orden", "diferencia_alas_vs_correcta"], key=lambda s: s if s.name == "_orden" else s.abs(), ascending=[True, False]).drop(columns="_orden")

    disputas = df[df["categoria"] == "alas_cobro_distinto"].copy()
    revisar = df[df["categoria"] == "revisar_manualmente"].copy()
    desactualizadas = df[df["categoria"] == "manualquote_desactualizada"].copy()
    sin_comuna = df[df["categoria"] == "comuna_no_reconocida"].copy()
    sin_dato_flapp = df[df["categoria"] == "sin_dato_flapp"].copy()
    sin_dato_alas = df[df["categoria"] == "sin_dato_alas"].copy()

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Hoja Detalle completo primero (para que las formulas de Resumen puedan apuntar a ella)
        df.to_excel(writer, sheet_name="Detalle completo", index=False)
        disputas.to_excel(writer, sheet_name="Disputas con Alas", index=False)
        revisar.to_excel(writer, sheet_name="Revisar manualmente", index=False)
        desactualizadas.to_excel(writer, sheet_name="ManualQuotes desact.", index=False)
        if len(sin_comuna):
            sin_comuna.to_excel(writer, sheet_name="Comunas no reconocidas", index=False)
        if len(sin_dato_flapp):
            sin_dato_flapp.to_excel(writer, sheet_name="Sin dato en Flapp", index=False)
        if len(sin_dato_alas):
            sin_dato_alas.to_excel(writer, sheet_name="Sin dato en Alas", index=False)

        wb = writer.book
        detalle_ws = wb["Detalle completo"]
        n_rows = len(df) + 1  # + header

        # ---- Hoja Resumen (con formulas SUMIFS/COUNTIFS sobre Detalle completo) ----
        resumen_ws = wb.create_sheet("Resumen Ejecutivo", 0)
        resumen_ws["A1"] = f"Conciliación de cobros Alas — {mes_label}"
        resumen_ws["A1"].font = Font(name="Arial", size=14, bold=True)
        resumen_ws["A2"] = "Generado automáticamente comparando: tarifa correcta según Matriz de Tarifas Flapp vs. lo cargado en Flapp vs. lo cobrado por Alas."
        resumen_ws["A2"].font = Font(name="Arial", size=9, italic=True)

        headers = ["Categoría", "Cantidad de envíos", "Suma diferencia (Alas - correcta) $", "Descripción"]
        desc = {
            "alas_cobro_distinto": "Flapp tenía la tarifa correcta cargada, pero Alas cobró un monto distinto. REQUIERE DISPUTA con Alas.",
            "revisar_manualmente": "Tanto lo cargado en Flapp como lo cobrado por Alas difieren de la tarifa correcta, y entre sí. Revisar caso a caso (posible cambio real de zona/bodega).",
            "manualquote_desactualizada": "Alas cobró correctamente, pero Flapp tenía cargada una tarifa vieja/errónea. No hay nada que reclamarle a Alas; actualizar ManualQuotes.",
            "sin_dato_flapp": "Alas facturó este envío pero no aparece en el export de Flapp. Revisar manualmente (¿otro courier, otro período?).",
            "sin_dato_alas": "El envío está en Flapp pero Alas no lo facturó en este archivo. Puede estar pendiente o fuera de período.",
            "comuna_no_reconocida": "No se pudo calcular la tarifa correcta (comuna no está en la matriz de tarifas). Revisar manualmente.",
            "cancelado_sin_cargo": "Envío cancelado, Alas no cobró. Informativo, no requiere acción.",
            "sin_discrepancia": "Todo cuadra: tarifa correcta = cargado en Flapp = cobrado por Alas.",
        }
        row0 = 4
        for j, h in enumerate(headers, start=1):
            c = resumen_ws.cell(row=row0, column=j, value=h)
            c.font = Font(name="Arial", bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="4472C4")
        # Verificar posición real de las columnas 'categoria' y 'diferencia_alas_vs_correcta'
        cols = list(df.columns)
        col_cat = get_column_letter(cols.index("categoria") + 1)
        col_dif = get_column_letter(cols.index("diferencia_alas_vs_correcta") + 1)

        r = row0
        for cat in orden_categorias:
            r += 1
            resumen_ws.cell(row=r, column=1, value=cat)
            resumen_ws.cell(row=r, column=2, value=f"=COUNTIF('Detalle completo'!{col_cat}2:{col_cat}{n_rows},A{r})")
            resumen_ws.cell(row=r, column=3, value=f"=SUMIF('Detalle completo'!{col_cat}2:{col_cat}{n_rows},A{r},'Detalle completo'!{col_dif}2:{col_dif}{n_rows})")
            resumen_ws.cell(row=r, column=4, value=desc.get(cat, ""))
        r += 1
        resumen_ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
        resumen_ws.cell(row=r, column=2, value=f"=SUM(B{row0+1}:B{r-1})").font = Font(bold=True)
        resumen_ws.cell(row=r, column=3, value=f"=SUM(C{row0+1}:C{r-1})").font = Font(bold=True)

        for col, width in zip("ABCD", (28, 20, 32, 90)):
            resumen_ws.column_dimensions[col].width = width
        for row in resumen_ws.iter_rows(min_row=row0+1, max_row=r):
            for cell in row:
                cell.font = Font(name="Arial", bold=cell.font.bold)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Formato de todas las hojas de datos: fuente Arial, encabezado en negrita, columnas auto
        for sheet_name in ["Detalle completo", "Disputas con Alas", "Revisar manualmente", "ManualQuotes desact.",
                           "Comunas no reconocidas", "Sin dato en Flapp", "Sin dato en Alas"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for cell in ws[1]:
                cell.font = Font(name="Arial", bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4472C4")
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(name="Arial")
            for i, col_cells in enumerate(ws.columns, start=1):
                max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
                ws.column_dimensions[get_column_letter(i)].width = min(max(10, max_len + 2), 45)

        wb.move_sheet("Resumen Ejecutivo", offset=-(len(wb.sheetnames)-1))

    return df, disputas, revisar, desactualizadas, sin_comuna


def construir_email(mes_label, disputas, revisar, desactualizadas, sin_comuna):
    total_disputa = disputas["diferencia_alas_vs_correcta"].sum() if len(disputas) else 0
    n_disputa = len(disputas)
    n_revisar = len(revisar)
    n_desact = len(desactualizadas)

    top = disputas.reindex(disputas["diferencia_alas_vs_correcta"].abs().sort_values(ascending=False).index).head(10)

    lineas_top = []
    for _, row in top.iterrows():
        signo = "cobró de más" if row["diferencia_alas_vs_correcta"] > 0 else "cobró de menos"
        lineas_top.append(
            f"  - Envío {row['shipmentId']} ({row['origen']} -> {row['destino']}, cliente {row['cliente']}): "
            f"tarifa correcta ${row['tarifa_correcta_matriz']:,.0f}, Alas cobró ${row['cobrado_por_alas']:,.0f} "
            f"({signo} ${abs(row['diferencia_alas_vs_correcta']):,.0f})"
        )
    lineas_top_txt = "\n".join(lineas_top) if lineas_top else "  (sin casos)"

    cuerpo = f"""Asunto: Conciliación cobros Alas — {mes_label} — {n_disputa} envíos en disputa

Hola,

Adjunto la conciliación de cobros de Alas correspondiente a {mes_label}. Comparamos, para cada
envío, la tarifa que debería aplicar según nuestra matriz de tarifas vigente, lo que teníamos
cargado en Flapp, y lo que Alas efectivamente cobró.

Resumen:
  - {n_disputa} envíos donde Alas cobró un monto distinto al que corresponde según la tarifa
    acordada (nuestra carga en Flapp era correcta). Diferencia neta: ${total_disputa:,.0f}
    ({"a favor nuestro" if total_disputa < 0 else "en contra nuestra" if total_disputa > 0 else "neutro"}).
  - {n_revisar} envíos requieren revisión manual (tanto lo cargado como lo cobrado difieren de
    la tarifa correcta, y entre sí — posible cambio real de zona o bodega ese mes).
  - {n_desact} envíos donde Alas cobró correctamente pero teníamos una tarifa vieja cargada en
    ManualQuotes (no es un problema de facturación de Alas; ya lo estamos corrigiendo
    internamente).

Los 10 casos de mayor impacto a disputar con Alas:
{lineas_top_txt}

Se adjunta el detalle completo en Excel (hoja "Disputas con Alas") con el desglose envío por
envío para que puedan revisarlo con Alas directamente.

Quedo atento a comentarios antes de enviarlo.

Saludos,
"""
    return cuerpo


def main():
    args = sys.argv[1:]
    alas_path = None
    if "--alas" in args:
        i = args.index("--alas")
        alas_path = args[i + 1]
        args = args[:i] + args[i + 2:]

    if len(args) < 2:
        print(__doc__)
        sys.exit(1)
    input_path = args[0]
    mes_label = args[1]
    out_dir = args[2] if len(args) > 2 else "."
    os.makedirs(out_dir, exist_ok=True)

    if alas_path:
        resultados = procesar_dos_archivos(input_path, alas_path)
    else:
        resultados = procesar(input_path)

    slug = mes_label.replace(" ", "_")
    xlsx_path = os.path.join(out_dir, f"Conciliacion_Alas_{slug}.xlsx")
    df, disputas, revisar, desactualizadas, sin_comuna = construir_excel(resultados, mes_label, xlsx_path)

    email_txt = construir_email(mes_label, disputas, revisar, desactualizadas, sin_comuna)
    email_path = os.path.join(out_dir, f"Borrador_Email_Alas_{slug}.md")
    with open(email_path, "w", encoding="utf-8") as f:
        f.write(email_txt)

    print(f"OK. Total filas procesadas: {len(resultados)}")
    print(Counter(r["categoria"] for r in resultados))
    print(f"Excel: {xlsx_path}")
    print(f"Email: {email_path}")


if __name__ == "__main__":
    main()
